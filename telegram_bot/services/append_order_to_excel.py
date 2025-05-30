import os
from datetime import datetime
from typing import List

from aiogram.types import ShippingAddress
from config import EXCEL_FILE, logger
from database.models import OrderItem
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


async def append_order_to_excel(
    user_id: int,
    shipping_address: ShippingAddress,
    order_items: List[OrderItem],
):
    logger.info(f"Начало сохранения заказа в Excel для user_id={user_id}")

    file_exists = os.path.exists(EXCEL_FILE)
    logger.debug(
        f"Проверка существования файла Excel: {EXCEL_FILE} — {file_exists}"
    )

    try:
        if file_exists:
            wb = load_workbook(EXCEL_FILE)
            ws: Worksheet = wb.active
            logger.debug("Файл Excel найден, загружаем рабочий лист")
        else:
            wb = Workbook()
            ws: Worksheet = wb.active
            ws.append(
                [
                    "Дата",
                    "User ID",
                    "Страна",
                    "Город",
                    "Улица",
                    "Индекс",
                    "Товары",
                    "Общая сумма (₽)",
                ]
            )
            logger.debug("Файл Excel не найден, создан новый файл и заголовки")

        # Формируем строку с товарами
        items_str = "\n".join(
            f"{item.product.name} (x{item.quantity}) — {item.product.price:.2f}₽"
            for item in order_items
            if item.product
        )
        total_price = sum(
            item.product.price * item.quantity
            for item in order_items
            if item.product
        )

        logger.debug(f"Сформирована строка с товарами: {items_str}")
        logger.debug(f"Общая сумма заказа: {total_price:.2f}₽")

        # Добавляем строку в Excel
        ws.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                user_id,
                shipping_address.country_code or "",
                shipping_address.city or "",
                shipping_address.street_line1 or "",
                shipping_address.post_code or "",
                items_str,
                round(total_price, 2),
            ]
        )

        wb.save(EXCEL_FILE)
        logger.info(f"Заказ пользователя {user_id} успешно сохранён в Excel")

    except Exception as e:
        logger.error(
            f"Ошибка при сохранении заказа в Excel для user_id={user_id}: {e}"
        )
        raise
